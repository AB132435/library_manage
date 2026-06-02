from flask import Blueprint, request, jsonify, send_file, current_app
from ..models import AuditLog
from .. import db
from flask_jwt_extended import jwt_required
from datetime import datetime, timedelta
from sqlalchemy import desc, func, cast, Date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import os

logs_bp = Blueprint('logs', __name__)


def add_log(username, module, action, detail=''):
    """添加审计日志的辅助函数"""
    try:
        log = AuditLog(
            username=username,
            module=module,
            action=action,
            detail=detail
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"[AuditLog Error] {e}")


@logs_bp.route('/', methods=['GET'])
@jwt_required()
def get_logs():
    """获取审计日志列表，支持筛选和分页"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    username = request.args.get('username', '')
    module = request.args.get('module', '')
    action = request.args.get('action', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = AuditLog.query

    if username:
        query = query.filter(AuditLog.username.contains(username))
    if module:
        query = query.filter_by(module=module)
    if action:
        query = query.filter_by(action=action)
    if date_from:
        try:
            dt = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AuditLog.op_time >= dt)
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, '%Y-%m-%d')
            query = query.filter(AuditLog.op_time < dt)
        except ValueError:
            pass

    query = query.order_by(desc(AuditLog.op_time))
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    logs = pagination.items

    return jsonify({
        "logs": [{
            "id": log.id,
            "op_time": log.op_time.isoformat() if log.op_time else None,
            "username": log.username,
            "module": log.module,
            "action": log.action,
            "detail": log.detail
        } for log in logs],
        "total": pagination.total,
        "page": page,
        "per_page": per_page,
        "pages": pagination.pages
    }), 200


@logs_bp.route('/modules', methods=['GET'])
@jwt_required()
def get_log_modules():
    """获取所有日志模块列表（用于筛选下拉框）"""
    modules = db.session.query(AuditLog.module).distinct().all()
    return jsonify([m[0] for m in modules if m[0]]), 200


@logs_bp.route('/stats', methods=['GET'])
@jwt_required()
def get_log_stats():
    """获取审计日志统计信息（用于看板）"""
    # 按模块统计
    module_stats = db.session.query(
        AuditLog.module,
        func.count(AuditLog.id).label('count')
    ).group_by(AuditLog.module).all()

    # 按操作类型统计
    action_stats = db.session.query(
        AuditLog.action,
        func.count(AuditLog.id).label('count')
    ).group_by(AuditLog.action).all()

    # 按日期统计（最近7天）- 使用 func.date 兼容 SQLite
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    date_stats = db.session.query(
        func.date(AuditLog.op_time).label('date'),
        func.count(AuditLog.id).label('count')
    ).filter(AuditLog.op_time >= seven_days_ago) \
     .group_by(func.date(AuditLog.op_time)) \
     .order_by(func.date(AuditLog.op_time)) \
     .all()

    # 今日操作数
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_count = AuditLog.query.filter(AuditLog.op_time >= today).count()

    # 总日志数
    total_count = AuditLog.query.count()

    return jsonify({
        "module_stats": [{"module": m[0], "count": m[1]} for m in module_stats],
        "action_stats": [{"action": a[0], "count": a[1]} for a in action_stats],
        "date_stats": [{"date": str(d[0]), "count": d[1]} for d in date_stats],
        "today_count": today_count,
        "total_count": total_count
    }), 200


@logs_bp.route('/export', methods=['GET'])
@jwt_required()
def export_logs():
    """导出审计日志为 Excel 文件"""
    username = request.args.get('username', '')
    module = request.args.get('module', '')
    action = request.args.get('action', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = AuditLog.query

    if username:
        query = query.filter(AuditLog.username.contains(username))
    if module:
        query = query.filter_by(module=module)
    if action:
        query = query.filter_by(action=action)
    if date_from:
        try:
            dt = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AuditLog.op_time >= dt)
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, '%Y-%m-%d')
            query = query.filter(AuditLog.op_time < dt)
        except ValueError:
            pass

    query = query.order_by(desc(AuditLog.op_time))
    logs = query.all()

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    # 设置表头样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = ['ID', '操作时间', '操作用户', '模块', '操作类型', '详细信息']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 填充数据
    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=log.id).border = thin_border
        ws.cell(row=row_idx, column=2, value=log.op_time.strftime('%Y-%m-%d %H:%M:%S') if log.op_time else '').border = thin_border
        ws.cell(row=row_idx, column=3, value=log.username).border = thin_border
        ws.cell(row=row_idx, column=4, value=log.module).border = thin_border
        ws.cell(row=row_idx, column=5, value=log.action).border = thin_border
        ws.cell(row=row_idx, column=6, value=log.detail or '').border = thin_border

    # 调整列宽
    column_widths = [10, 20, 15, 15, 15, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 保存到内存缓冲区
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'audit_logs_{timestamp}.xlsx'

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@logs_bp.route('/import', methods=['POST'])
@jwt_required()
def import_logs():
    """从 Excel 文件导入审计日志"""
    if 'file' not in request.files:
        return jsonify({"error": "未找到上传文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "未选择文件"}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({"error": "仅支持 .xlsx 格式的 Excel 文件"}), 400

    try:
        # 读取 Excel 文件
        wb = load_workbook(filename=BytesIO(file.read()))
        ws = wb.active

        # 跳过表头，从第二行开始读取
        imported_count = 0
        skipped_count = 0
        error_messages = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if all(cell is None for cell in row):
                continue  # 跳过空行

            try:
                # 解析数据：ID, 操作时间，操作用户，模块，操作类型，详细信息
                log_id, op_time_str, username, module, action, detail = row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else ''

                # 验证必填字段
                if not username or not module or not action:
                    error_messages.append(f"第{row_idx}行：缺少必填字段（用户名、模块或操作类型）")
                    skipped_count += 1
                    continue

                # 解析时间
                op_time = None
                if op_time_str:
                    if isinstance(op_time_str, datetime):
                        op_time = op_time_str
                    else:
                        try:
                            # 尝试多种时间格式
                            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d']:
                                try:
                                    op_time = datetime.strptime(str(op_time_str), fmt)
                                    break
                                except ValueError:
                                    continue
                            if op_time is None:
                                raise ValueError("无法解析的时间格式")
                        except Exception:
                            op_time = datetime.utcnow()  # 使用当前时间作为默认值

                # 创建日志记录
                log = AuditLog(
                    username=str(username),
                    module=str(module),
                    action=str(action),
                    detail=str(detail) if detail else '',
                    op_time=op_time
                )
                db.session.add(log)
                imported_count += 1

            except Exception as e:
                error_messages.append(f"第{row_idx}行：{str(e)}")
                skipped_count += 1
                continue

        db.session.commit()

        result = {
            "message": "导入完成",
            "success_count": imported_count,
            "skipped_count": skipped_count
        }
        if error_messages:
            result["errors"] = error_messages[:10]  # 最多返回 10 条错误信息

        return jsonify(result), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"导入失败：{str(e)}"}), 500

